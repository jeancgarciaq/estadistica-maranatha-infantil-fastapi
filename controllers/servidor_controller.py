import logging
from datetime import datetime
from models.servidor import Servidor
from sqlalchemy import extract
from sqlalchemy.exc import SQLAlchemyError
from controllers import BaseController

# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ServidorController(BaseController):
    def __init__(self, session=None):
        super().__init__(model=Servidor, session=session)
        logger.info("ServidorController inicializado.")

    def crear_servidor(self, datos, user_context=None):
        """
        Crea un nuevo registro de servidor en la base de datos.
        :param datos: Diccionario con los datos del servidor.
        :return: (Exito, Mensaje)
        """
        if not isinstance(datos, dict):
            return False, "Los datos proporcionados no son válidos."

        datos_normalizados = self._normalizar_datos(datos)

        def operacion(db):
            errores = self.validar_datos(datos_normalizados, db, is_new=True)
            if errores:
                raise ValueError("\n".join(errores))

            servidor = Servidor(**datos_normalizados)
            db.add(servidor)
            db.flush()
            self.registrar_evento_sync(db, 'servidores', servidor, 'upsert')
            logger.info("Servidor creado.")

        return self.ejecutar_transaccion(operacion, "Servidor creado exitosamente.", user_context=user_context)

    def listar_servidores(self, filtros=None):
        """
        Lista todos los registros de servidores desde la base de datos.
        :param filtros: Diccionario con criterios de búsqueda.
        :return: Lista de objetos Servidor.
        """
        db = self.get_db_session()
        try:
            query = self.query_activa(db)
            if filtros:
                if filtros.get('nombre'):
                    query = query.filter(Servidor.nombre.ilike(f"%{filtros['nombre']}%"))
                if filtros.get('cedula'):
                    query = query.filter(Servidor.cedula == int(filtros['cedula']))
                if filtros.get('celular'):
                    query = query.filter(Servidor.celular.ilike(f"%{filtros['celular']}%"))
                if filtros.get('correo'):
                    query = query.filter(Servidor.correo.ilike(f"%{filtros['correo']}%"))
                if filtros.get('area_servicio'):
                    query = query.filter(Servidor.area_servicio.ilike(f"%{filtros['area_servicio']}%"))
                if filtros.get('mes_nacimiento'):
                    query = query.filter(extract('month', Servidor.fecha_nacimiento) == int(filtros['mes_nacimiento']))
                if filtros.get('dia_nacimiento'):
                    query = query.filter(extract('day', Servidor.fecha_nacimiento) == int(filtros['dia_nacimiento']))

            servidores = query.order_by(Servidor.nombre.asc()).all()
            logger.info(f"{len(servidores)} servidores obtenidos.")
            return servidores
        except SQLAlchemyError as e:
            logger.error(f"Error al listar servidores: {e}")
            return []
        finally:
            if not self.session:
                db.close()

    def obtener_servidor(self, id):
        """
        Obtiene un registro de servidor por su ID.
        :param id: ID del servidor.
        :return: Objeto Servidor o None.
        """
        db = self.get_db_session()
        try:
            return self.query_activa(db).filter(Servidor.id == id).first()
        except SQLAlchemyError as e:
            logger.error(f"Error al obtener servidor: {e}")
            return None
        finally:
            if not self.session:
                db.close()

    def actualizar_servidor(self, id, datos, user_context=None):
        """
        Actualiza un registro de servidor existente.
        :param id: ID del servidor a actualizar.
        :param datos: Diccionario con los datos actualizados.
        :return: (Exito, Mensaje)
        """
        if not id or not isinstance(id, int):
            return False, "El ID del servidor es obligatorio y debe ser un número entero."

        datos_normalizados = self._normalizar_datos(datos)

        def operacion(db):
            servidor = db.query(Servidor).filter(Servidor.id == id, Servidor.is_deleted.is_(False)).first()
            if not servidor:
                raise ValueError("Servidor no encontrado.")

            # Combinar datos existentes con los nuevos para la validación
            datos_para_validar = {k: getattr(servidor, k) for k in servidor.__table__.columns.keys()}
            datos_para_validar.update(datos_normalizados)

            errores = self.validar_datos(datos_para_validar, db, is_new=False, current_id=id)
            if errores:
                raise ValueError("\n".join(errores))

            for key, value in datos_normalizados.items():
                setattr(servidor, key, value)
            
            self.registrar_evento_sync(db, 'servidores', servidor, 'upsert')
            logger.info(f"Servidor actualizado: ID {id}")

        return self.ejecutar_transaccion(operacion, "Servidor actualizado exitosamente.", user_context=user_context)

    def generar_reporte_excel(self, servidores):
        """Genera un archivo Excel (.xlsx) con la lista de servidores."""
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista de Servidores"

        # Encabezados
        headers = ["Nombre", "Cédula", "Edad", "F. Nacimiento", "Celular", "Correo", "Equipo", "Área", "Capitán"]
        ws.append(headers)

        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for s in servidores:
            ws.append([
                s.nombre,
                s.cedula,
                s.edad,
                str(s.fecha_nacimiento) if s.fecha_nacimiento else "N/A",
                s.celular or "N/A",
                s.correo or "N/A",
                s.numero_equipo or 0,
                s.area_servicio or "N/A",
                s.capitan or "N/A"
            ])

        # Ajuste de ancho de columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def generar_reporte_pdf(self, servidores):
        """Genera un archivo PDF con la lista de servidores."""
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Lista de Personal de Servicio (Servidores)", styles['Title']))
        elements.append(Spacer(1, 12))

        data = [["Nombre", "Cédula", "Edad", "Celular", "Correo", "Área de Servicio", "Capitán"]]
        for s in servidores:
            data.append([
                s.nombre,
                str(s.cedula),
                str(s.edad),
                s.celular or "N/A",
                s.correo or "N/A",
                s.area_servicio or "N/A",
                s.capitan or "N/A"
            ])

        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(t)
        
        doc.build(elements)
        pdf_value = buffer.getvalue()
        buffer.close()
        return pdf_value

    def eliminar_servidor(self, id, user_context=None):
        """
        Elimina un registro de servidor por su ID.
        :param id: ID del servidor a eliminar.
        :return: (Exito, Mensaje)
        """
        if not id or not isinstance(id, int):
            return False, "El ID del servidor es obligatorio y debe ser un número entero."

        def operacion(db):
            servidor = db.query(Servidor).filter(Servidor.id == id, Servidor.is_deleted.is_(False)).first()
            if not servidor:
                raise ValueError("Servidor no encontrado.")
            
            self.marcar_eliminado(servidor, db)
            self.registrar_evento_sync(db, 'servidores', servidor, 'delete')
            logger.info(f"Servidor eliminado: ID {id}")

        return self.ejecutar_transaccion(operacion, "Servidor eliminado exitosamente.", user_context=user_context)

    def _normalizar_datos(self, datos):
        """
        Normaliza los datos de entrada para facilitar validaciones.
        """
        datos_normalizados = dict(datos)
        for campo in ["nombre", "celular", "correo", "area_servicio", "capitan"]:
            if campo in datos_normalizados and isinstance(datos_normalizados[campo], str):
                datos_normalizados[campo] = datos_normalizados[campo].strip()
                if datos_normalizados[campo] == "":
                    datos_normalizados[campo] = None
        
        # Convertir campos numéricos que puedan venir como string vacío a None
        for campo in ["edad", "cedula", "numero_equipo"]:
            if campo in datos_normalizados and datos_normalizados[campo] == "":
                datos_normalizados[campo] = None

        if "fecha_nacimiento" in datos_normalizados and not datos_normalizados["fecha_nacimiento"]:
             datos_normalizados["fecha_nacimiento"] = None

        return datos_normalizados

    def validar_datos(self, datos, db, is_new=True, current_id=None):
        """
        Valida los datos proporcionados para crear o actualizar un servidor.
        :param datos: Diccionario con los datos del servidor.
        :param db: Sesión de la base de datos.
        :param is_new: Booleano que indica si es una creación (True) o actualización (False).
        :param current_id: ID del servidor actual si es una actualización.
        :return: Lista de errores encontrados.
        """
        errores = []

        nombre = datos.get("nombre")
        if not nombre or not isinstance(nombre, str) or not nombre.strip():
            errores.append("El campo 'nombre' es obligatorio y debe ser una cadena de texto.")
        elif len(nombre) > 100:
            errores.append("El campo 'nombre' no puede superar los 100 caracteres.")

        edad = datos.get("edad")
        if edad is None:
            errores.append("El campo 'edad' es obligatorio.")
        elif not isinstance(edad, int) or edad <= 0:
            errores.append("El campo 'edad' debe ser un número entero positivo.")

        cedula = datos.get("cedula")
        if cedula is None:
            errores.append("El campo 'cedula' es obligatorio.")
        elif not isinstance(cedula, int) or cedula <= 0:
            errores.append("El campo 'cedula' debe ser un número entero positivo.")
        else:
            query = db.query(Servidor).filter(Servidor.cedula == cedula, Servidor.is_deleted.is_(False))
            if not is_new and current_id:
                query = query.filter(Servidor.id != current_id)
            if query.first():
                errores.append(f"Ya existe un servidor con la cédula {cedula}.")

        fecha_nacimiento = datos.get("fecha_nacimiento")
        if fecha_nacimiento:
            try:
                if isinstance(fecha_nacimiento, str):
                    datetime.strptime(fecha_nacimiento, '%Y-%m-%d')
            except ValueError:
                errores.append("El campo 'fecha de nacimiento' debe tener el formato 'YYYY-MM-DD'.")

        celular = datos.get("celular")
        if celular and (not isinstance(celular, str) or len(celular) > 20):
            errores.append("El campo 'celular' debe ser una cadena de texto de hasta 20 caracteres.")

        correo = datos.get("correo")
        if correo:
            if not isinstance(correo, str) or len(correo) > 100 or "@" not in correo:
                errores.append("El campo 'correo' debe ser una dirección de correo electrónico válida de hasta 100 caracteres.")
            else:
                query = db.query(Servidor).filter(Servidor.correo == correo, Servidor.is_deleted.is_(False))
                if not is_new and current_id:
                    query = query.filter(Servidor.id != current_id)
                if query.first():
                    errores.append(f"Ya existe un servidor con el correo electrónico {correo}.")

        numero_equipo = datos.get("numero_equipo")
        if numero_equipo is not None and (not isinstance(numero_equipo, int) or numero_equipo <= 0):
            errores.append("El campo 'numero de equipo' debe ser un número entero positivo.")

        area_servicio = datos.get("area_servicio")
        if area_servicio and (not isinstance(area_servicio, str) or len(area_servicio) > 100):
            errores.append("El campo 'area de servicio' debe ser una cadena de texto de hasta 100 caracteres.")

        capitan = datos.get("capitan")
        if capitan and (not isinstance(capitan, str) or len(capitan) > 100):
            errores.append("El campo 'capitan' debe ser una cadena de texto de hasta 100 caracteres.")

        return errores